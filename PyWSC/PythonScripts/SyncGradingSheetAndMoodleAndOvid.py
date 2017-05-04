if __name__ == '__main__':
    import os, sys

    sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), '..'))
    # sys.path.append(os.path.join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))), '..'))

from openpyxl import cell as Cell
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from PythonScripts.Imports import Config
from PythonScripts.Imports import OvidOverSSH as ossh
from PythonScripts.Imports import Utility as util
from PythonScripts.Imports import WebService as ws
from PythonScripts.Imports.Classes import Grade

"""
    Script to synchronize an excel sheet with a moodle cource as configured in config.ini.
    This script is rather long and unhandy, so it is tiled in various sections.
        1. Read in worksheet.
        2. Read in moodle course.
        3. Compare both and present differences to user. If the user is OK with changes to be made - continue.
        4. Do necessary actions to synchronize moodle with sheet. F.e. enrol students and perform group switches.
           Group changes are made simultaneously on OVID over SSH, if enabled in config.ini.
        5. Upload gradings (and comments if enabled).

    If it finds any assignment ID, group name or matrikel number not matching with moodle, it will stop with a
    hopefully helpful message to the user.
    TODO: proper exception handling to correct these values on the fly...
    TODO: refactoring and modularizing

"""

def get_group_id_by_name(group_name, student_name):
    group_id= -1
    result_get_course_groups = ws.do_post_on_webservice('core_group_get_course_groups', {'courseid': current_course_id})
    for group in result_get_course_groups:
        if group['name'] == group_name:
            group_id = group['id']
            break

    if group_id == -1:
        util.print_error_message_and_prompt_any_key_to_exit(
            'Gruppen ID in Moodle für Wechsel nicht gefunden!' + '\n'
            + 'Student: ' + student_name  + '\n'
            + 'Gruppe: ' + group_name
        )
    return group_id


def get_comment_value_by_assignment_id(marked_grade_cell: Cell, marked_comment_cells: Cell, row_number, ws: worksheet):
    """
        Helper method to find the corresponding comment value for a given grade_cell.
        Column index marked by assignment ID in upload_marker row. Row indexing by given row_number.
        The list of cells (marked_comment_cells) is cleaned before and do not contain the
        CommentMarkerPrefix anymore, but assignment IDs.

    :param marked_grade_cell: Grade cell marked by assignment ID in the upload_marker row.
    :param marked_comment_cells: List of all comment cells we want to find the assignment ID of marked_grade_cell in.
    :param row_number: Number of current row.
    :param ws: Worksheet to read from (data_only).
    :return: Empty string if cell value is None or if cell is not found.
    """
    if Config.get_behavior_comments_switch():
        for coment_cell in marked_comment_cells:
            # find matching assignment id
            if str(coment_cell.value) == str(marked_grade_cell.value):
                # extract coordinates (column index)
                column_index_of_comment_cell = column_index_from_string(
                    coordinate_from_string(coment_cell.coordinate)[0])
                # read value of cell
                comment_value = ws.cell(row=row_number, column=column_index_of_comment_cell).value
                if not comment_value: # normalize - we want string only as return type
                    comment_value = ''
                return comment_value

    if Config.get_behavior_log():
        print('Comment for grade cell ' + str(marked_grade_cell.coordinate) + ' not found!')
    return ''


util.ask_user_to_continue('Wollen Sie wirklich Moodle mit ihrer Tabelle synchronisieren?\n'
                          + 'Tabelle: ' + Config.get_grading_sheet_file_path() + '\n'
                          + 'Fortfahren? (ja/j)')

# Open workbook in data_only mode to read values only and no formulas.
# Workbook in data_only mode will return calculated values for formulas: value == internal_value
wb_grading_data = load_workbook(Config.get_grading_sheet_file_path(), data_only=True)
ws_grading_data = wb_grading_data.active

# find upload_marker row
moodle_grade_column_marker_name = Config.get_column_mapping_upload_marker_name()
index_moodle_grade_marker_row = util.find_row_of_value_in_column(moodle_grade_column_marker_name,
                                                                 Config.get_column_mapping_upload_marker_column(),
                                                                 ws_grading_data)

# switch to enable grading including comments to accomplish advanced point systems to match with moodle simple point system
use_grading_comments = Config.get_behavior_comments_switch()

# Get column indexes as defined in config.ini
index_matrikel_col = column_index_from_string(Config.get_column_mapping_matrnr_column())
index_group_name_col = column_index_from_string(Config.get_column_mapping_group_name_column())

#default_email_domain = Config.get_email_default_email_domain_students()
current_course_id = Config.get_current_course_id()

# **********************************************************************************
# READ IN MARKED MOODLE GRADE COLUMNS (and comment columns), STUDENTS AND GROUP NAMES
# **********************************************************************************

# Get all moodle grade columns marked with assignment IDs by upload_marker row as configured in config.ini
# First: get all not None value cells
moodle_grade_column_marker_cells = util.get_cells_of_not_none_value_in_row(index_moodle_grade_marker_row,
                                                                           ws_grading_data)

# Then: clean list - we want marked moodle grade columns only
# Also: extract moodle grade comment cells
comment_marker_prefix = Config.get_column_mapping_comment_marker_prefix()
grade_comment_cells = []

for cell in moodle_grade_column_marker_cells:
    # Remove cell that contain upload_marker row name
    if cell.value == moodle_grade_column_marker_name:
        moodle_grade_column_marker_cells.remove(cell)
    # If enabled: remove moodle grade comment cells and add them to list grade_comment_cells
    elif use_grading_comments:
        # Found a moodle grade comment cell
        if str(comment_marker_prefix) in str(cell.value):
            # Clean cell value by removing marker prefix
            cell.value = str(cell.value).strip(str(comment_marker_prefix))
            grade_comment_cells.append(cell)
            moodle_grade_column_marker_cells.remove(cell)

# Get all cells containing matrikel numbers in matrikel column as configured in config.ini
# First: get all not None value cells
student_matrikel_number_cells = util.get_cells_of_not_none_value_in_column(index_matrikel_col, ws_grading_data)

# Then: clean list
# We want matrikel numbers only
for cell in student_matrikel_number_cells:
    if not str(cell.value).isdigit():
        student_matrikel_number_cells.remove(cell)

# We want to ignore certain rows, as configured in config.ini
# in section [GRADING_SHEET_COLUMN_MAPPING][IgnoreGradingOnEntryOfValue]
#  --- HOTFIX(IgnoredStudents)
#for cell in student_matrikel_number_cells:
 #   current_cell_row = coordinate_from_string(cell.coordinate)[1]

    #if util.check_for_values_in_row(Config.get_column_mapping_ignore_grading_on_entry(), current_cell_row, ws_grading_data):
        #student_matrikel_number_cells.remove(cell)

# ************************************************************************************
# READ IN DATA FROM SHEET
# ************************************************************************************

students_in_front_sheet = []

for matrnr_cell in student_matrikel_number_cells:

    # current_row = 42 <- ('A', 42) <- 'A42'
    current_row = coordinate_from_string(matrnr_cell.coordinate)[1]

    current_matrnr = matrnr_cell.value

    # get group name, we maybe have to search it
    group_name = ws_grading_data.cell(row=current_row, column=index_group_name_col).value

    if not group_name:
        util.print_error_message_and_prompt_any_key_to_exit(
            'Gruppenname in Zeile ' + str(
                current_row) + ' nicht gefunden.\nDer Gruppenname muss in jeder Zeile stehen, in der auch Matrikelnummern vorkommen.'
        )

    current_student = {'username': str(current_matrnr),
                       # 'firstname': ws_front_data.cell(row=current_row, column=index_first_name_col).value,
                       # 'lastname': ws_front_data.cell(row=current_row, column=index_last_name_col).value,
                       # 'email': str(current_matrikel_number) + default_email_domain,
                       'group': {
                           'new': {
                               'name': group_name,
                               'id': ''},
                           'old': {
                               'name': '',
                               'id': ''}},
                       'id': '',
                       'grades': []
                       }
    for grade_marker_cell in moodle_grade_column_marker_cells:
        # current_grade_col_index = 1 <- 'A' <- ('A', 42) <- 'A42'
        current_grade_col_index = column_index_from_string(coordinate_from_string(grade_marker_cell.coordinate)[0])

        current_grade_value = ws_grading_data.cell(row=current_row, column=current_grade_col_index).value

        # ignore None value grades
        if current_grade_value:
            current_assignment_id = grade_marker_cell.value

            current_student['grades'].append(
                Grade.Grade(current_grade_value,
                            get_comment_value_by_assignment_id(grade_marker_cell, grade_comment_cells, current_row,
                                                               ws_grading_data),
                            'id_not_set',
                            current_assignment_id))

    students_in_front_sheet.append(current_student)

students_in_front_sheet = sorted(students_in_front_sheet, key=lambda user: user['username'])

# ************************************************************************************
# READ IN MOODLE COURSE
# ************************************************************************************

result_get_enrolled_users = ws.do_post_on_webservice('core_enrol_get_enrolled_users', {'courseid': current_course_id})
result_get_course_groups = ws.do_post_on_webservice('core_group_get_course_groups', {'courseid': current_course_id})

enrolled_users_in_course = []
for enrolled_user in result_get_enrolled_users:
    group_details = {}
    if enrolled_user['groups']:  # enrolled user is in a group
        group_details = {
            'name': enrolled_user['groups'][0]['name'],
            'id': str(enrolled_user['groups'][0]['id'])}

    enrolled_users_in_course.append(
        {'username': str(enrolled_user['username']),
         # 'firstname': enrolled_user['firstname'],
         # 'lastname': enrolled_user['lastname'],
         # 'email': enrolled_user['email'],
         'group': group_details,  # {} for enrolled user without a group
         'id': str(enrolled_user['id'])})

# clean lists - we dont want the PyWsClient in the lists
for user in enrolled_users_in_course:
    if util.check_username_is_client_name(user[Config.get_moodle_user_id_field_name()]):
        enrolled_users_in_course.remove(user)
    elif 'roles' in user.keys() and user['roles']:
        # we do not want any other user than STUDENT role in list
        if str(user['roles']['roleid']) != str(Config.get_enrolment_role_id_student()):
            enrolled_users_in_course.remove(user)

# check for valid assignment IDs in sheet
# first: get all assignments from moodle course
result_get_assignments = ws.do_post_on_webservice('mod_assign_get_assignments', {'courseids[0]': current_course_id})

# then: compare assignment IDs in sheet with IDs in moodle
assignment_id_is_valid = False
for assignment in result_get_assignments['courses'][0]['assignments']:
    for cell in moodle_grade_column_marker_cells:
        if str(assignment['id']) == str(cell.value):
            assignment_id_is_valid = True

    if not assignment_id_is_valid:
        # stop script here
        util.print_error_message_and_prompt_any_key_to_exit(
            'Assignment IDs in Excel stimmen nicht mit den von Moodle überein!')

if use_grading_comments:
    # then: compare comment cell assignment IDs in sheet with IDs in moodle
    assignment_id_is_valid = False
    for assignment in result_get_assignments['courses'][0]['assignments']:
        for cell in grade_comment_cells:
            if str(assignment['id']) == str(cell.value):
                assignment_id_is_valid = True

        if not assignment_id_is_valid:
            # stop script here
            util.print_error_message_and_prompt_any_key_to_exit(
                'Assignment IDs in Excel stimmen nicht mit denen von Moodle überein!')

# ************************************************************************************
# COMPARE USER LISTS
# ************************************************************************************

# sort list
enrolled_users_in_course = sorted(enrolled_users_in_course, key=lambda user: user['username'])

for user_in_sheet in students_in_front_sheet:
    for user_in_moodle in enrolled_users_in_course:

        # find user_in_sheet in enrolled_users_in_course
        if user_in_sheet['username'] == user_in_moodle['username']:

            # set moodle id for user_in_sheet
            user_in_sheet['id'] = user_in_moodle['id']

            # set moodle id for user_in_sheet for each grade (easier for upload)
            for grade in user_in_sheet['grades']:
                grade.user_id = user_in_moodle['id']

            # find desired group in moodle course
            desired_group_id = -1
            for group in result_get_course_groups:
                if group['name'] == user_in_sheet['group']['new']['name']:
                    desired_group_id = group['id']
            if desired_group_id > -1:  # group name found
                # set moodle group ID to student_in_sheet
                user_in_sheet['group']['new']['id'] = str(desired_group_id)
            else:  # group name not found - exit
                util.print_error_message_and_prompt_any_key_to_exit(
                    'Gruppen ID in Moodle für Wechsel nicht gefunden!' + '\n'
                    + 'Student: ' + user_in_sheet['username']
                    + ', Gruppe: ' + user_in_sheet['group']['new']['name']
                )

            # check if group switch is necessary
            # user_in_moodle is enrolled, and has a group
            if user_in_moodle['group']:
                # save old group name and id
                user_in_sheet['group']['old']['name'] = user_in_moodle['group']['name']
                user_in_sheet['group']['old']['id'] = user_in_moodle['group']['id']

# ************************************************************************************
# PREVIEW CHANGES TO USER
# ************************************************************************************

any_changes_neccessary = False
preview_string = ''
for student in students_in_front_sheet:
    enroll_string = ''
    group_switch_string = ''

    if not student['id']:
        enroll_string = 'Muss eingeschrieben werden.\n'
        any_changes_neccessary = True

    if str(student['group']['old']['name']) != str(student['group']['new']['name']):
        any_changes_neccessary = True
        group_switch_string = (student['username'] + ':\n'
                               + enroll_string
                               + 'Gruppenwechsel: '
                               + str(student['group']['old']['name']) + ' -> ' + str(
            student['group']['new']['name'])) + '\n'

    if any_changes_neccessary:
        preview_string += group_switch_string

if any_changes_neccessary:
    print('************************\n' +
          '     PREVIEW CHANGES\n' +
          '************************')
    print(preview_string)
    util.ask_user_to_continue()

# ************************************************************************************
# ENROLL UNENROLLED USERS
# ************************************************************************************

if any_changes_neccessary:
    students_to_enrol = []
    for student in students_in_front_sheet:

        if not student['id']:
            # get user information from moodle
            result_get_user = ws.do_post_on_webservice(
                'core_user_get_users_by_field',
                {'field': Config.get_moodle_user_id_field_name(),
                 'values[0]': student['username']})

            # if student not in Moodle, result is empty
            if not result_get_user:
                util.print_error_message_and_prompt_any_key_to_exit(
                    'Student nicht in Moodle registriert!' + '\n'
                    + 'Student: ' + student['username'] + '\n'
                    + str(result_get_user)
                )

            student['id'] = result_get_user[0]['id']
            students_to_enrol.append(student['id'])
    if students_to_enrol:
        ws.enrol_students(students_to_enrol)

# ************************************************************************************
# DO GROUP CHANGES
# ************************************************************************************

if any_changes_neccessary:
    students_to_delete_group_membership = []
    students_to_add_group_membership = []
    students_that_need_new_password = []
    students_that_keep_old_password = []

    for student in students_in_front_sheet:

        if not student['group']['old']['id']:
            students_to_add_group_membership.append(student)
            students_that_need_new_password.append(student)
        elif student['group']['old']['id'] != student['group']['new']['id']:
            students_to_add_group_membership.append(student)
            students_to_delete_group_membership.append(student)
            students_that_keep_old_password.append(student)

    if students_to_delete_group_membership:
        payload_delete_group_members = {}
        payload_delete_counter = 0
        for student in students_to_delete_group_membership:
            payload_delete_group_members.update({
                'members[' + str(payload_delete_counter) + '][groupid]': student['group']['old']['id'],
                'members[' + str(payload_delete_counter) + '][userid]': student['id']
            })
            payload_delete_counter += 1

        ws.do_post_on_webservice('core_group_delete_group_members', payload_delete_group_members)

    if students_to_add_group_membership:
        payload_add_group_members = {}
        payload_add_counter = 0
        for student in students_to_add_group_membership:

            # hotfix: if student was not enrolled we are missing new group id here - this script should be reworked :(
            if not student['group']['new']['id']:
                student['group']['new']['id'] = get_group_id_by_name(student['group']['new']['name'], student['username'])

            payload_add_group_members.update({
                'members[' + str(payload_add_counter) + '][groupid]': student['group']['new']['id'],
                'members[' + str(payload_add_counter) + '][userid]': student['id']
            })
            payload_add_counter += 1

        ws.do_post_on_webservice('core_group_add_group_members', payload_add_group_members)

# ************************************************************************************
# DO GROUP CHANGES IN OVID
# ************************************************************************************

if any_changes_neccessary:
    if Config.get_behavior_send_commands_to_ovid():
        util.ask_user_to_continue('Zur Kommunikation mit OVID wird VPN benötigt!\nBitte Starten Sie VPN!\nFortfahren? (ja/nein)')
        for student in students_that_need_new_password:
            ossh.write_new_pw_to_group_and_send_email_to_student(student['username'], student['group']['new']['name'])

        for student in students_that_keep_old_password:
            ossh.write_pw_from_old_group_to_new_group_and_send_email_to_student(student['username'],
                                                                                student['group']['old']['name'],
                                                                                student['group']['new']['name'])

# ************************************************************************************
# UPLOAD GRADES but SORT OUR STUDENTS TO IGNORE GRADINGS
#  --- HOTFIX(IgnoredStudents): vorher wurden alle Studenten, in deren Zeilen Werte aus IgnoreGradingOnEntryOfValue standen, auch für Gruppenwechsel ignoriert...
#
# We want to ignore certain rows, as configured in config.ini
# in section [GRADING_SHEET_COLUMN_MAPPING][IgnoreGradingOnEntryOfValue]
# ************************************************************************************
for student in students_in_front_sheet:
    for cell in student_matrikel_number_cells:
        if str(cell.value) == student['username']:
            current_cell_row = coordinate_from_string(cell.coordinate)[1]
            if not util.check_for_values_in_row(current_cell_row,
                                        ws_grading_data):
                ws.upload_grades(student['grades'])

# ************************************************************************************
# UPLOAD GRADES
# ************************************************************************************
#  --- HOTFIX(IgnoredStudents)
#for student in students_in_front_sheet:
   # ws.upload_grades(student['grades'])

# open course overview in moodle
util.open_moodle_in_webbrowser('course/view.php?id=' + str(current_course_id))
