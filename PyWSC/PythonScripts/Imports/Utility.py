#!/usr/bin/env python
import json
import os.path
import subprocess
import sys
import webbrowser

from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.styles import numbers

from PythonScripts.Imports import Config

"""
    Util package for this client. Holds useful, essential, often used methods.
"""

PRESS_ANY_KEY_MESSAGE = 'Beliebiege Taste zum Beenden Drücken'

def print_json_pretty(text):
    """
        Prints json text with indent to console to make it readable for humans.
        Used to print responses of moodle.
    :param text: text in json formt
    :return: None
    """
    if text:
        json_text = json.dumps(text, indent=4, sort_keys=True)
        if (json_text):
            print(json_text)
    else:
        print('[]')


def adjust_workbook_format(wb: Workbook):
    """
        Beautifies a Workbook by adjusting the width of each column to widest cell.value in that column.
        Also sets the format of cells to TEXT format. This disables the automatically conversion of cell
        values by Excel.
    :param wb: Workbook that shall be displayed properly and nicely in f.e. excel.
    :return: None
    """

    # iterate over each worksheet
    ws_names = wb.get_sheet_names()
    for n in range(0, len(ws_names)):
        ws = wb.get_sheet_by_name(ws_names[n])

        # iterate over each column of worksheet
        for column in ws.columns:
            max_col_width = 0

            # iterate over each cell in column
            for cell in column:

                # set format of cell as plain text
                cell.number_format = numbers.FORMAT_TEXT

                if cell.value:  # not empty
                    # check cell.value length
                    if len(str(cell.value)) > max_col_width:
                        # set as new column_width
                        max_col_width = len(str(cell.value)) + 2

            # only set column_width if column is not empty
            if max_col_width > 0:
                ws.column_dimensions[column[0].column].width = max_col_width


def find_row_of_value_in_column(value_to_find, column_letter, ws: worksheet):
    """
        Finds a value in column with column_letter in worksheet ws.

    :param value_to_find: value we want to find
    :param column_letter: letter of column in ws, f.e. 'A' 'B' 'C' ...
    :param ws: worksheet in workbook
    :return: the index of the row we found the value in - or None
    """
    for i in range(1, ws.max_row + 1):
        if ws[column_letter + str(i)].value == value_to_find:
            return i


def check_for_values_in_row(row_number, ws: worksheet):
    """
        Checks for any occurrences of value(s) in given row in given worksheet as defined in
        [GRADING_SHEET_COLUMN_MAPPING][IgnoreGradingOnEntryOfValue]

    :param row_number: Number of row to check.
    :param ws: Worksheet(data_only mode!) to check row from.
    :return: True if value(s) are in row - otherwise False.
    """
    values = Config.get_column_mapping_ignore_grading_on_entry()
    for i in range(1, ws.max_column + 1):
        if any(val in str(ws.cell(row=row_number, column=i).value) for val in values):
            return True
    return False


def get_cells_of_not_none_value_in_row(row_number, ws: worksheet):
    cell_list = []
    full_row = ws.iter_rows(min_row=row_number, max_col=ws.max_column, max_row=row_number)
    for cells in full_row:
        for n in range(0, len(cells)):
            if cells[n].value:
                cell_list.append(cells[n])
    return cell_list


def get_cells_of_not_none_value_in_column(col_number, ws: worksheet):
    cell_list = []
    full_col = ws.iter_rows(min_col=col_number, max_row=ws.max_row, max_col=col_number)
    for cells in full_col:
        for n in range(0, len(cells)):
            if cells[n].value:
                cell_list.append(cells[n])
    return cell_list


def open_moodle_in_webbrowser(parameter_string, open_in_new_tab=2):
    """
        Opens the moodle website as defined in config.ini in [BAHAVIOR] according to operating system
        with default webbrowser.

    :param parameter_string: additional url parameters, concat with moodle domain address as configured in config.ini.
    :param open_in_new_tab: 2 is default (defined by package webbrowser)
    :return: None
    """
    if Config.get_behavior_moodle():
        # open_in_new_tab = 2 means: open in a new tab, if possible
        url = Config.get_moodle_domain() + parameter_string
        webbrowser.open(url, new=open_in_new_tab)


def open_workbook(absolut_path):
    """
        Opens the grading workbook as defined in config.ini in [BAHAVIOR] according to operating system
        with default program for viewing *.xlsx-files. Should support Windows, MacOS and Linux.

    :return: None
    """
    if Config.get_behavior_excel():
        if sys.platform.startswith('darwin'):  # Mac OS
            subprocess.call(('open', absolut_path))
        elif os.name == 'nt':  # Windows
            os.startfile(absolut_path)
        elif os.name == 'posix':  # Linux
            subprocess.call(('xdg-open', absolut_path))


def wait_for_user_input_to_keep_console_opened():
    """
        As configured in config.ini in [BAHAVIOR] the user will be prompted to press any key to finish the script and exit the console.
        To be used f.e. when logging to console.

    :return: None
    """
    if Config.get_behavior_user_prompt():
        input(PRESS_ANY_KEY_MESSAGE)


def ask_user_to_continue(question_text='Fortfahren? (ja/nein)', answer_to_continue=('j', 'ja')):
    decision = input(
        question_text).lower() in answer_to_continue
    if not decision:
        sys.exit('Abbruch durch Benutzer.')
    return decision


def ask_user_to_overwrite_file(file_path):
    if os.path.isfile(file_path):
        do_overwrite = input(
            "Datei existiert: "
            + file_path
            + "\nSoll diese überschrieben werden? (ja/nein)").lower() in ('j', 'ja')
        if not do_overwrite:
            sys.exit('Abbruch durch Benutzer.')


def print_error_message_and_prompt_any_key_to_exit(error_message,
                                                   prompt_any_key_message=PRESS_ANY_KEY_MESSAGE):
    print('******************** FEHLER ********************')
    print(str(error_message) + '\n')
    press_any_key_to_exit(prompt_any_key_message)


def press_any_key_to_exit(message=PRESS_ANY_KEY_MESSAGE):
    input(message)
    sys.exit('')


def check_for_student_role(id):
    return str(id) == str(Config.get_enrolment_role_id_student())


def check_username_is_client_name(username):
    return str(username) == str(Config.get_py_client_name())
