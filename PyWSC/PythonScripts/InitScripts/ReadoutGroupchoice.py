#!/usr/bin/env python
"""
    Script to create an excel sheet containing group names and group members (matrikel numbers) for
    further initiation of OVID. Resulting sheet is in following format:
        A       B       C       D       ...
    1   1-1     12345   123916  12313
    2   1-2     13140   13157
    ...

"""
import sys, os
from openpyxl import Workbook

from PythonScripts.Imports import Config
from PythonScripts.Imports import Utility as util
from PythonScripts.Imports import WebService as ws
from PythonScripts.Imports.Classes import Group

# ask user if ['OUTPUT_FILE']['OutputFileNameStudent2Group'] shall be overwritten
util.ask_user_to_overwrite_file(Config.get_output_student2group())

# create a workbook in memory
wb_out = Workbook()
ws_out = wb_out.active

# get all groups of current moodle course
result_get_groups = ws.do_post_on_webservice('core_group_get_course_groups',
                                             {'courseid': Config.get_current_course_id()})

# create list of Group for simpler handling
groups = []
for group in result_get_groups:
    groups.append(Group.Group(group['id'], group['name'], []))

# sort for proper displaying in excel
groups = sorted(groups, key=lambda group: group.id)

# create payload for api call
group_counter = 0
payload = {}
for group in groups:
    payload.update({'groupids[' + str(group_counter) + ']': group.id})
    group_counter += 1

# get all group members
result_get_group_members = ws.do_post_on_webservice('core_group_get_group_members', payload)

for group_with_members in result_get_group_members:
    # filter empty groups - we only want groups containing members
    if group_with_members['userids'] and len(group_with_members['userids']) > 0:
        # find corresponding group in Group list
        for group in groups:
            if str(group.id) == str(group_with_members['groupid']):
                # get proper user information (matrikel number) for each member
                for member in group_with_members['userids']:
                    result_get_users = ws.do_post_on_webservice('core_user_get_users', {
                        'criteria[0][key]': 'id',
                        'criteria[0][value]': member  # just user_id
                    })
                    # fill members of group in Group list with matrikel number
                    group.add_member(result_get_users['users'][0][Config.get_moodle_user_id_field_name()])
                break

# write Group list to spreadsheet
row_counter = 1
for group in groups:
    if group.has_members():
        # write group name in first column
        ws_out.cell(row=row_counter, column=1).value = group.name
        column_counter = 2
        for member in group.members:
            # write matrikel numbers of members next to corresponding group name
            ws_out.cell(row=row_counter, column=column_counter).value = str(member)
            column_counter += 1
        row_counter += 1

# beautify workbook
util.adjust_workbook_format(wb_out)

# save workbook
wb_out.save(Config.get_output_student2group())

# open workbook in excel
util.open_workbook(Config.get_output_student2group())

# wait for user prompt
util.wait_for_user_input_to_keep_console_opened()

if __name__ == '__main__':
    import os, sys

    sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), '..'))
    sys.path.append(os.path.join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))), '..'))
