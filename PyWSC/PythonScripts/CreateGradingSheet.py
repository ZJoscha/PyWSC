#!/usr/bin/env python
if __name__ == '__main__':
    import os, sys

    sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), '..'))
    #sys.path.append(os.path.join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))), '..'))

import os
import sys

from openpyxl import Workbook

from PythonScripts.Imports import Config
from PythonScripts.Imports import Utility as util
from PythonScripts.Imports import WebService as ws
from PythonScripts.Imports.Classes import Group
from PythonScripts.Imports.Classes import Student

"""
    Creates Workbook containing user, grade, group and assignment information of the moodle course as defined in
    config.ini in section [CURRENT_COURSE]. Also the Upload_Marker row is written, to make the resulting workbook
    suitable for synchronization with the moodle course (by SyncGradingSheetAndMoodle.py).

    The resulting Workbook will contain four worksheets:
        - 'grades' :        Contains an uploadable or copy&pastable sheet to upload grades manually into moodle. Contains essential data
                            of students and columns for each assignment to enable grading and feedback input.
        - 'assignments' :   Contains all gradable assignments (Übung) in moodle course. To upload grades programmatically we need
                            the assignment_IDs which are held in this sheet (name and ID).
        - 'groups' :        Contains all groups of current moodle course (just name and ID).
        - 'not_in_group':   Contains all
"""

# ask user if ['OUTPUT_FILE']['OutputFileNameGradings'] shall be overwritten
util.ask_user_to_overwrite_file(Config.get_output_grading_sheet_file_path())

# initialise Workbook and Worksheets
wb_out = Workbook()
ws_grades = wb_out.active  # Workbook is created with empty sheet
ws_grades.title = 'grades'  # rename that empty sheet
ws_assignments = wb_out.create_sheet('assignments')
ws_groups = wb_out.create_sheet('groups')
ws_not_in_group = wb_out.create_sheet('not_in_group')

course_id = Config.get_current_course_id()
moodle_uid = Config.get_moodle_user_id_field_name()

# ****************************************************************
# START OF API CALLS - START TO FILL WORKBOOK
# ****************************************************************

# get groups of course
result_get_course_groups = ws.do_post_on_webservice('core_group_get_course_groups', {'courseid': course_id})

# get enrolled users
result_get_enrolled_users = ws.do_post_on_webservice('core_enrol_get_enrolled_users', {'courseid': course_id})

list_of_enrolled_users_in_group = []
list_of_enrolled_users_not_in_group = []

groups = []
for grp in result_get_course_groups:
    if grp:
        groups.append(Group.Group(grp['id'], grp['name'], []))

# sort for proper display in excel
groups = sorted(groups, key=lambda group: group.id)  # re.findall("([\d])-([\d])",group.name))

# ****************************************************************
# WRITE GROUPS TO WORKBOOK
# ****************************************************************

# write groups to worksheet
ws_groups['A1'] = 'group_name'
ws_groups['B1'] = 'group_id'

row_counter = 2
for grp in groups:
    ws_groups['A' + str(row_counter)] = grp.name
    ws_groups['B' + str(row_counter)] = grp.id
    row_counter += 1

# ****************************************************************
# GET GROUPS AND MEMBERS
# ****************************************************************
# get members of each group
for n in range(0, len(groups)):
    result_get_group_members = ws.do_post_on_webservice('core_group_get_group_members', {'groupids[0]': groups[n].id})

    # get user information for each member of groups
    for i in range(0, len(result_get_group_members)):
        for userid in result_get_group_members[i]['userids']:
            # get user information
            result_get_user = ws.do_post_on_webservice('core_user_get_users',
                                                       {'criteria[0][key]': 'id',
                                                        'criteria[0][value]': userid})

            # put member in group
            groups[n].add_member(
                Student.Student(result_get_user['users'][0]['username'],
                        userid,
                        result_get_user['users'][0].get('email',result_get_user['users'][0]['username']+Config.get_email_default_email_domain_students()),
                        result_get_user['users'][0]['firstname'],
                        result_get_user['users'][0]['lastname'],
                        groups[n].name))

            # hold enrolled students in a group in separate list
            list_of_enrolled_users_in_group.append(userid)

# find enrolled users without group
for enrolled_user in result_get_enrolled_users:
    has_group = False

    # compare enrolled users in a group to enrolled users
    for enrolled_user_in_group in list_of_enrolled_users_in_group:
        if str(enrolled_user['id']) == str(enrolled_user_in_group):
            has_group = True

    if not has_group:
        # hold enrolled users not in group in separate list
        if 'roles' in enrolled_user.keys():
            if enrolled_user['roles']:
                # only add students to this list
                if util.check_for_student_role(enrolled_user['roles'][0]['roleid']) \
                        and not util.check_username_is_client_name(enrolled_user[Config.get_moodle_user_id_field_name()]):
                    list_of_enrolled_users_not_in_group.append(enrolled_user)

# ****************************************************************
# WRITE STUDENTS WITH NO GROUP TO WORKBOOK
# ****************************************************************

# prepare headers for no_group worksheet
header = [
    'matrikel_nr',
    'first_name',
    'last_name',
    'user_email',
    'user_id',
]

# write header to grading worksheet
for i in range(0, len(header)):
    ws_not_in_group.cell(row=1, column=i + 1).value = header[i]

# write students not in group to no_group worksheet
row_counter = 2
for student in list_of_enrolled_users_not_in_group:
    ws_not_in_group['A' + str(row_counter)] = student[moodle_uid]
    ws_not_in_group['B' + str(row_counter)] = student['firstname']
    ws_not_in_group['C' + str(row_counter)] = student['lastname']
    ws_not_in_group['D' + str(row_counter)] = student['email']
    ws_not_in_group['E' + str(row_counter)] = student['id']
    row_counter += 1

# ****************************************************************
# GET ASSIGNMENTS AND WRITE THEM TO WORKBOOK
# ****************************************************************

# get assignments of course
result_get_assignments = ws.do_post_on_webservice('mod_assign_get_assignments', {'courseids[0]': course_id})

# if result['courses'] and result['courses'][0] and 'assignments' in result['courses'][0]:
assignments = []
for assignment in result_get_assignments['courses'][0]['assignments']:
    assignments.append({'id': assignment['id'], 'name': assignment['name']})

# write assignments to assignments worksheet
ws_assignments['A1'] = 'assignment_id'
ws_assignments['B1'] = 'assignment_name'

for i in range(0, len(assignments)):
    ws_assignments['A' + str(i + 2)] = assignments[i]['id']
    ws_assignments['B' + str(i + 2)] = assignments[i]['name']

# ****************************************************************
# WRITE GRADING WORKSHEET
# ****************************************************************

# prepare headers for grading worksheet
header = [
    'group_name',
    'matrikel_nr',
    'first_name',
    'last_name',
    'user_email',
    'user_id',
    'group_id'
]

# in case you modify the basic fields of user information, we save the start value here
assignment_columns_start_index = len(header) + 1

# add to header: assignment and feedback columns
for assignment in assignments:
    header.append(assignment['name'])
    header.append(assignment['name'] + ' (feedback)')

header.append('Total')
header.append('Total (feedback)')

# write header to grading worksheet
for i in range(0, len(header)):
    ws_grades.cell(row=1, column=i + 1).value = header[i]

# write students to grading worksheet
row_counter = 2

for group in groups:
    for member in group.members:
        ws_grades['A' + str(row_counter)] = group.name
        ws_grades['B' + str(row_counter)] = member.name
        ws_grades['C' + str(row_counter)] = member.firstname
        ws_grades['D' + str(row_counter)] = member.lastname
        ws_grades['E' + str(row_counter)] = member.email
        ws_grades['F' + str(row_counter)] = member.id
        ws_grades['G' + str(row_counter)] = group.id

        # get grades of current student
        result_get_grades = ws.do_post_on_webservice('core_grades_get_grades',{
            'courseid':course_id,
            'userids[0]': member.id
        })
        #assignment_columns_range = range(8, (2 * len(assignments)) +1)
        assignment_column_counter = assignment_columns_start_index
        total_column = assignment_column_counter + (2 * len(assignments))

        # handle first item in result_get_grades different - it is the course total
        ws_grades.cell(row=row_counter, column=total_column).value = result_get_grades['items'][0]['grades'][0]['grade']
        ws_grades.cell(row=row_counter, column=total_column + 1).value = result_get_grades['items'][0]['grades'][0][
            'str_feedback']

        assignment_item_counter =1

        for assignment in assignments:
                ws_grades.cell(row=row_counter,column=assignment_column_counter).value = \
                    result_get_grades['items'][assignment_item_counter]['grades'][0]['grade']
                ws_grades.cell(row=row_counter, column=assignment_column_counter +1).value = \
                    result_get_grades['items'][assignment_item_counter]['grades'][0]['str_feedback']
                assignment_item_counter += 1
                assignment_column_counter += 2

        row_counter += 1

# write grading markers
# add one blank line
row_counter+=1
# write UploadMarkerRowName
ws_grades.cell(row=row_counter, column=1).value = str(Config.get_column_mapping_upload_marker_name())

# write assignment IDs and comment marker prefixes into grading (and feedback) columns
column_counter = 8
for assignment in assignments:
    ws_grades.cell(row=row_counter, column=column_counter).value = str(assignment['id'])
    if Config.get_behavior_comments_switch():
        ws_grades.cell(row=row_counter, column=column_counter+1).value = str(
            Config.get_column_mapping_comment_marker_prefix()) + str(assignment['id'])
    column_counter += 2

# ****************************************************************
# END OF SCRIPT - BEAUTIFY AND SAVE WORKBOOK
# ****************************************************************

# beautify workbook
util.adjust_workbook_format(wb_out)

# save workbook in directory configured in config.ini
wb_out.save(Config.get_output_grading_sheet_file_path())

# open workbook in default program
util.open_workbook(Config.get_output_grading_sheet_file_path())

# wait for user prompt
util.wait_for_user_input_to_keep_console_opened()
